# -*- coding: utf-8 -*-
import bridge, os, re, tempfile, time, threading
from pathlib import Path
from datetime import datetime
import pandas as pd
import pdfplumber
import fitz
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

os.environ["DISABLE_MODEL_SOURCE_CHECK"] = "True"
try:
    from rapidocr_onnxruntime import RapidOCR
    ENGINE = RapidOCR()
except Exception:
    ENGINE = None


def _start_heartbeat(label, state, interval=5):
    stop_event = threading.Event()
    start_time = time.time()

    def worker():
        while not stop_event.wait(interval):
            stage = state.get("stage", "\u5904\u7406\u4e2d")
            current = state.get("current")
            total = state.get("total")
            elapsed = int(time.time() - start_time)
            if current is not None and total:
                bridge.update_terminal(f"\u23f3 [{label}] {stage}: {current}/{total}\uff0c\u5df2\u8017\u65f6 {elapsed}s")
            else:
                bridge.update_terminal(f"\u23f3 [{label}] {stage}\uff0c\u5df2\u8017\u65f6 {elapsed}s")

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()
    return stop_event, thread


def _stop_heartbeat(stop_event, thread):
    stop_event.set()
    try:
        thread.join(timeout=1)
    except Exception:
        pass


def clean_text(text: str) -> str:
    if not text: return ""
    return text.replace(" ", "").replace("　", "").replace("：", ":").replace("，", ",").replace("¥", "").replace("￥", "").replace("_", "").replace("~", "")

def get_standard_rate(ocr_rate_str: str):
    if not ocr_rate_str: return None
    nums = re.findall(r"\d+", ocr_rate_str)
    if not nums:
        if "免" in ocr_rate_str: return 0.0
        return None
    val = int(nums[0])
    if val in [13, 9, 6, 3, 1, 0]: return val / 100.0
    if val % 10 == 1: return 0.01
    if val % 10 == 3: return 0.03
    if val % 10 == 6: return 0.06
    if str(val).endswith("13"): return 0.13
    return None

def calculate_rate_from_amounts(pre_tax: float, tax: float):
    if not pre_tax: return None
    actual_rate = tax / pre_tax
    standards = [0.13, 0.09, 0.06, 0.03, 0.01, 0.00]
    best_rate, min_diff = 0.01, 1.0
    for r in standards:
        diff = abs(actual_rate - r)
        if diff < min_diff: min_diff, best_rate = diff, r
    return best_rate

def to_float(s):
    try: return float(s) if s not in (None, "") else None
    except Exception: return None

def parse_date(date_str: str):
    if not date_str: return None
    try:
        match = re.search(r"(20\d{2})[^\d]*(\d{1,2})[^\d]*(\d{1,2})", date_str)
        if match:
            y, m, d = map(int, match.groups())
            return datetime(y, m, d)
    except Exception: return None
    return None

def parse_rate(rate_str: str):
    if not rate_str: return None
    try:
        if "免" in rate_str: return 0.0
        return float(rate_str.replace("%", "")) / 100.0
    except Exception: return None

def ocr_image_file(image_path: str, progress_callback=None) -> str:
    if ENGINE is None:
        return ""
    try:
        if progress_callback:
            progress_callback("OCR \u56fe\u50cf\u8bc6\u522b")
        result, _ = ENGINE(image_path)
        if result:
            if progress_callback:
                progress_callback("OCR \u56fe\u50cf\u8bc6\u522b")
            return " ".join([line[1] for line in result])
    except Exception:
        return ""
    return ""

def find_exact_pair_sum(numbers, target, tolerance=0.03):
    candidates = sorted([n for n in numbers if 0.05 < n < target], reverse=True)
    for i in range(len(candidates)):
        val_a = candidates[i]
        target_b = target - val_a
        for j in range(len(candidates)):
            if i == j: continue
            val_b = candidates[j]
            if abs(val_b - target_b) <= tolerance:
                return max(val_a, val_b), min(val_a, val_b)
    return None, None

def find_subset_sum(numbers, target, tolerance=0.05):
    candidates = [n for n in numbers if 0.5 < n <= target + tolerance]
    for n in candidates:
        if abs(n - target) <= tolerance: return n
    candidates = sorted(candidates, reverse=True)[:15]
    for i in range(len(candidates)):
        for j in range(i + 1, len(candidates)):
            s = candidates[i] + candidates[j]
            if abs(s - target) <= tolerance: return s
    for i in range(len(candidates)):
        for j in range(i + 1, len(candidates)):
            for k in range(j + 1, len(candidates)):
                s = candidates[i] + candidates[j] + candidates[k]
                if abs(s - target) <= tolerance: return s
    return None

def get_info_strict_code(full_text: str):
    clean_str = clean_text(full_text)
    
    invoice_code = ""
    code_match = re.search(r"(?<!信用)代码[^0-9]{0,8}(\d{10,12})(?!\d)", clean_str)
    if code_match:
        temp_code = code_match.group(1)
        if not temp_code.startswith("91"): invoice_code = temp_code

    invoice_num = ""
    num_match = re.search(r"号码[^0-9]{0,10}(\d{8,20})", clean_str)
    if num_match: invoice_num = num_match.group(1)
    if not invoice_num:
        idx = clean_str.find("发票")
        window = clean_str[max(0, idx - 50): idx + 200] if idx != -1 else clean_str
        long_num_match = re.search(r"(?<!\d)(\d{20})(?!\d)", window)
        if long_num_match: invoice_num = long_num_match.group(1)

    invoice_date = ""
    date_match = re.search(r"(20\d{2}[^0-9a-zA-Z]{1,3}\d{1,2}[^0-9a-zA-Z]{1,3}\d{1,2})", clean_str)
    if date_match: invoice_date = date_match.group(1)

    all_nums = []
    for n in re.findall(r"([0-9]+\.\d{2})", clean_str):
        try:
            f = float(n)
            if 0.01 <= f < 10000000000: all_nums.append(f)
        except: continue
    floats = sorted(all_nums, reverse=True)

    ocr_rate, ocr_rate_str = 0.01, "1%"
    rate_match = re.search(r"(税率|征收率)[^%免0-9]{0,15}(\d{1,2}%|免税)", clean_str)
    if rate_match:
        cleaned = get_standard_rate(rate_match.group(2))
        if cleaned is not None:
            ocr_rate = cleaned
            ocr_rate_str = f"{int(ocr_rate * 100)}%" if ocr_rate > 0 else "免税"
    else:
        rate_match2 = re.search(r"(\d{1,3}%|免税)", clean_str)
        if rate_match2:
            cleaned = get_standard_rate(rate_match2.group(0))
            if cleaned is not None:
                ocr_rate = cleaned
                ocr_rate_str = f"{int(ocr_rate * 100)}%" if ocr_rate > 0 else "免税"

    total_val = 0.0
    total_match = re.search(r"(?:价税合计|小写)[^0-9]{0,15}([0-9,]+\.\d{2})", clean_str)
    if total_match:
        try: total_val = float(total_match.group(1).replace(",", ""))
        except: pass
    if total_val == 0.0 and floats: total_val = floats[0]

    pre_tax_val = tax_val = None
    tax_match = re.search(r"(?:合计税额|税额)[^0-9]{0,15}([0-9,]+\.\d{2})", clean_str)
    if tax_match:
        try: tax_val = float(tax_match.group(1).replace(",", ""))
        except: pass
    amt_match = re.search(r"(?:不含税金额|合计金额|金额)[^0-9]{0,15}([0-9,]+\.\d{2})", clean_str)
    if amt_match:
        try: pre_tax_val = float(amt_match.group(1).replace(",", ""))
        except: pass

    final_pre_tax, final_tax = 0.0, 0.0
    final_rate_str = ocr_rate_str

    if total_val > 0:
        if pre_tax_val is not None and tax_val is not None:
            if abs((pre_tax_val + tax_val) - total_val) <= 0.05:
                final_pre_tax, final_tax = pre_tax_val, tax_val
                calc_rate = calculate_rate_from_amounts(final_pre_tax, final_tax)
                if calc_rate is not None:
                    final_rate_str = f"{int(calc_rate * 100)}%" if calc_rate > 0 else "免税"
            else:
                pre_tax_val = tax_val = None

        if pre_tax_val is None or tax_val is None:
            found_pre, found_tax = find_exact_pair_sum(floats, total_val)
            if found_pre:
                final_pre_tax, final_tax = found_pre, found_tax
                calc_rate = calculate_rate_from_amounts(final_pre_tax, final_tax)
                if calc_rate is not None:
                    final_rate_str = f"{int(calc_rate * 100)}%" if calc_rate > 0 else "免税"
            else:
                target_pre_tax = total_val / (1 + ocr_rate) if (1 + ocr_rate) != 0 else 0.0
                found_sum = find_subset_sum(floats, target_pre_tax)
                if found_sum:
                    final_pre_tax = found_sum
                    final_tax = total_val - final_pre_tax
                else:
                    final_pre_tax = target_pre_tax
                    final_tax = total_val - final_pre_tax

    def fmt(val): return "{:.2f}".format(val) if val and val > 0 else ""
    return (invoice_code, invoice_num, invoice_date, final_rate_str, fmt(final_pre_tax), fmt(final_tax), fmt(total_val))

def extract_text_from_pdf(file_path: str, progress_callback=None) -> str:
    text_parts = []
    try:
        with pdfplumber.open(file_path) as pdf:
            total = min(len(pdf.pages), 3)
            for i in range(total):
                if progress_callback:
                    progress_callback("PDF \u6587\u672c\u63d0\u53d6", i + 1, total)
                t = pdf.pages[i].extract_text() or ""
                if t:
                    text_parts.append(t)
    except Exception:
        pass

    full_text = "\n".join(text_parts)
    if len(clean_text(full_text)) >= 30:
        return full_text

    if ENGINE is None:
        return full_text
    try:
        with fitz.open(file_path) as doc:
            ocr_parts = []
            total = min(doc.page_count, 2)
            for i in range(total):
                if progress_callback:
                    progress_callback("PDF OCR \u56de\u9000", i + 1, total)
                page = doc[i]
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                    temp_img = tmp.name
                try:
                    pix.save(temp_img)
                    ocr_parts.append(ocr_image_file(temp_img, lambda stage: progress_callback(stage, i + 1, total) if progress_callback else None))
                finally:
                    if os.path.exists(temp_img):
                        os.remove(temp_img)
            ocr_text = "\n".join([p for p in ocr_parts if p])
            return ocr_text if len(clean_text(ocr_text)) > len(clean_text(full_text)) else full_text
    except Exception:
        return full_text

def save_excel_formatted(df: pd.DataFrame, output_file: str):
    df.to_excel(output_file, index=False)
    wb = load_workbook(output_file)
    ws = wb.active
    ws.title = "汇总"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    for col, w in {"A": 16, "B": 26, "C": 16, "D": 8, "E": 16, "F": 12, "G": 14}.items():
        ws.column_dimensions[col].width = w

    header_font = Font(name="微软雅黑", bold=True, size=12)
    body_font = Font(name="微软雅黑", size=11)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    border = Border(left=Side(style="thin", color="999999"), right=Side(style="thin", color="999999"), 
                    top=Side(style="thin", color="999999"), bottom=Side(style="thin", color="999999"))

    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.font = header_font; cell.alignment = center; cell.border = border

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20
        ws[f"A{r}"].number_format = "@"; ws[f"B{r}"].number_format = "@"; ws[f"C{r}"].number_format = 'yyyy"年"mm"月"dd"日"'
        ws[f"D{r}"].number_format = "0%"
        for col in ["E", "F", "G"]: ws[f"{col}{r}"].number_format = "#,##0.00"
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            cell = ws[f"{col}{r}"]
            cell.font = body_font; cell.border = border
            cell.alignment = right if col in ["E", "F", "G"] else center
    wb.save(output_file)

@bridge.expose
def run_invoice(path, recursive):
    try:
        folder = Path(path).resolve()
        files = list(folder.rglob("*.*") if recursive else folder.iterdir())
        valid = [f for f in files if f.suffix.lower() in (".pdf",".jpg",".png",".jpeg",".bmp") and not f.name.startswith("~$")]
        
        data_list, error_list, seen_invoices = [], [], {}
        for idx, f in enumerate(valid, 1):
            bridge.update_terminal(f"[*] \u6b63\u5728\u5904\u7406\u53d1\u7968 ({idx}/{len(valid)}): {f.name}")
            ext = f.suffix.lower().lstrip(".")
            state = {"stage": "\u63d0\u53d6\u6587\u672c", "current": idx, "total": len(valid)}
            stop_event, heartbeat_thread = _start_heartbeat(f.name, state)
            try:
                def progress(stage, current=None, total=None):
                    state["stage"] = stage
                    state["current"] = current if current is not None else idx
                    state["total"] = total if total is not None else len(valid)

                full_text = extract_text_from_pdf(str(f), progress) if ext == "pdf" else ocr_image_file(str(f), lambda stage: progress(stage))
                state.update({"stage": "\u89e3\u6790\u53d1\u7968\u5b57\u6bb5", "current": idx, "total": len(valid)})
            finally:
                _stop_heartbeat(stop_event, heartbeat_thread)
            
            if not full_text or len(clean_text(full_text)) < 10:
                error_list.append(f"{f.name}\t提取文本失败")
                continue

            code, num, date_str, rate_str, pre_tax_str, tax_str, total_str = get_info_strict_code(full_text)
            code_s, num_s = (str(code) if code else "").strip(), (str(num) if num else "").strip()
            
            if num_s and (code_s, num_s) in seen_invoices:
                error_list.append(f"{f.name}\t与 {seen_invoices[(code_s, num_s)]} 重复，已跳过")
                continue
            
            if num_s: seen_invoices[(code_s, num_s)] = f.name
            
            data_list.append({
                "发票代码": code_s, "发票号码": num_s, "开票日期": parse_date(date_str),
                "税率": parse_rate(rate_str), "金额(不含税)": to_float(pre_tax_str),
                "税额": to_float(tax_str), "价税合计": to_float(total_str)
            })
            
        if not data_list: return {"status": "error", "msg": "未提取到任何有效数据"}
        
        out_file = folder / "发票汇总表.xlsx"
        df_out = pd.DataFrame(data_list)[["发票代码", "发票号码", "开票日期", "税率", "金额(不含税)", "税额", "价税合计"]]
        bridge.update_terminal("[*] \u6b63\u5728\u5199\u5165\u53d1\u7968\u6c47\u603b Excel...")
        save_excel_formatted(df_out, str(out_file))
        
        msg = f"成功提取 {len(data_list)} 张"
        if error_list:
            (folder / "errors.txt").write_text("\n".join(error_list), encoding="utf-8")
            msg += f"，跳过/异常 {len(error_list)} 个 (详见 errors.txt)"
            
        return {"status": "success", "msg": msg}
    except Exception as e:
        return {"status": "error", "msg": str(e)}
